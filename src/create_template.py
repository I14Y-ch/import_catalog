import pandas as pd
import os
import requests
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Protection, Alignment
from openpyxl.worksheet.protection import SheetProtection
from datetime import datetime

# Define the columns for the template
columns = [
    # Mandatory fields
    'title',
    'description',
    'identificator',
    'accessRights',

    'issued', 
    'modified',
    
    # Keywords (three entries)
    'keywords_1',
    'keywords_2',
    'keywords_3',
    
    # Recommended fields
    'contactPoints_fn',
    'contactPoints_hasEmail',
    'contactPoints_hasTelephone',
    
    'themes_label',
    
    'spatial',
    
    'temporalCoverage_start',
    'temporalCoverage_end',
        
    # Distribution (recommended)
    'distribution_accessUrl_1', 
    'distribution_downloadUrl_1',
    'distribution_license_label_1',
    'distribution_accessUrl_2',
    'distribution_downloadUrl_2',
    'distribution_license_label_2',
    'distribution_accessUrl_3',
    'distribution_downloadUrl_3',
    'distribution_license_label_3',
]

def is_url_column(column_name):
    return 'url' in column_name.lower()

def is_date_column(column_name):
    date_columns = ['issued', 'modified', 'temporalCoverage_start', 'temporalCoverage_end']
    return column_name in date_columns

# Fetch themes codelist
def get_themes_codelist():
    url = "https://api.i14y.admin.ch/api/public/v1/concepts/08da58dc-4dc8-f9cb-b6f2-7d16b3fa0cde/codelist-entries/exports/json"
    response = requests.get(url)
    data = response.json()
    
    return pd.DataFrame([
        {'code': item.get('code', ''), 'label': item.get('name', {}).get('de', '')}
        for item in data['data']
    ])

def get_license_codelist():
    url = "https://api.i14y.admin.ch/api/public/v1/concepts/08db7eb9-8d92-b301-982e-5f7cbd44e45f/codelist-entries/exports/json"
    response = requests.get(url)
    data = response.json()
    
    licenses = [{'code': 'UNKNOWN', 'label': 'Unknown'}]
    licenses.extend([
        {'code': item.get('code', ''), 'label': item.get('name', {}).get('de', '')}
        for item in data['data']
    ])
    
    return pd.DataFrame(licenses)

def get_access_rights():
    access_rights = [
        {'code': 'NON_PUBLIC', 'label': 'Nicht-öffentlich'},
        {'code': 'PUBLIC', 'label': 'Öffentlich'},
        {'code': 'RESTRICTED', 'label': 'Eingeschränkt'},
        {'code': 'CONFIDENTIAL', 'label': 'Vertraulich'}
    ]
    return pd.DataFrame(access_rights)

# Create the main template DataFrame
df = pd.DataFrame(columns=columns)
df.loc[0] = ''

# Get themes codelist
themes_df = get_themes_codelist()

# Get license codelist
license_df = get_license_codelist()

# Get access rights codelist
access_rights_df = get_access_rights()

# Create data directory if it doesn't exist
os.makedirs('data', exist_ok=True)

# Define required DCAT fields
required_fields = ['title', 'description', 'identificator', 'accessRights']

# Define column widths
column_widths = {
    'title': 30,
    'description': 50,
    'distribution_license_label_1': 30,
    'distribution_license_label_2': 30,
    'distribution_license_label_3': 30,
}
DEFAULT_WIDTH = 20

# Create Excel writer object
with pd.ExcelWriter('data/template_datasets.xlsx', engine='openpyxl') as writer:
    # Write main template
    df.to_excel(writer, sheet_name='Template', index=False)
    
    # Write reference sheets
    themes_df.to_excel(writer, sheet_name='ThemesCodes', index=False)
    license_df.to_excel(writer, sheet_name='LicenseCodes', index=False)
    access_rights_df.to_excel(writer, sheet_name='AccessRights', index=False)
    
    workbook = writer.book
    template_sheet = workbook['Template']
    
    # Format header row
    grey_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    bold_font = Font(bold=True)
    normal_font = Font(bold=False)
    
     # Remove full sheet protection, we'll only protect the header row
    template_sheet.protection.sheet = False
    
    # Format headers and protect first row
    template_sheet.row_dimensions[1].height = 45
    
    for col_idx, col_name in enumerate(columns):
        cell = template_sheet.cell(row=1, column=col_idx + 1)
        cell.fill = grey_fill
        cell.protection = Protection(locked=True)
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        
        if col_name in required_fields:
            cell.font = bold_font
        else:
            cell.font = normal_font
    
    # Ensure all data cells are unprotected
    for row in template_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.protection = Protection(locked=False)
    
    # Adjust column widths
    for col_idx, col_name in enumerate(columns):
        column_letter = chr(65 + col_idx)
        width = column_widths.get(col_name, DEFAULT_WIDTH)
        template_sheet.column_dimensions[column_letter].width = width
        template_sheet.column_dimensions[column_letter].bestFit = True
    
    # Set row height and enable text wrapping
    for row in template_sheet.iter_rows(min_row=1, max_row=1000):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center')
    
    for row_idx in range(2, 1001):
        template_sheet.row_dimensions[row_idx].height = 30
    
    # Add validations
    themes_dv = DataValidation(
        type="list",
        formula1=f"=ThemesCodes!$B$2:$B${len(themes_df) + 1}",
        allow_blank=True
    )
    template_sheet.add_data_validation(themes_dv)
    themes_label_idx = columns.index('themes_label')
    for row in range(2, 1000):
        themes_dv.add(f"{chr(65 + themes_label_idx)}{row}")
    
    license_dv = DataValidation(
        type="list",
        formula1=f"=LicenseCodes!$B$2:$B${len(license_df) + 1}",
        allow_blank=True
    )
    template_sheet.add_data_validation(license_dv)
    for i in range(1, 4):
        col_idx = columns.index(f'distribution_license_label_{i}')
        for row in range(2, 1000):
            license_dv.add(f"{chr(65 + col_idx)}{row}")
    
    access_rights_dv = DataValidation(
        type="list",
        formula1=f"=AccessRights!$B$2:$B${len(access_rights_df) + 1}",
        allow_blank=True
    )
    template_sheet.add_data_validation(access_rights_dv)
    access_rights_idx = columns.index('accessRights')
    for row in range(2, 1000):
        access_rights_dv.add(f"{chr(65 + access_rights_idx)}{row}")

    # Apply date formatting
    for col_name in columns:
        if is_date_column(col_name):
            col_idx = columns.index(col_name)
            col_letter = chr(65 + col_idx)
            for row in range(2, 1000):
                cell = template_sheet[f"{col_letter}{row}"]
                cell.number_format = 'YYYY-MM-DD'
